"""Build Gold_Money_Data.xlsx: all data + model results in one workbook."""
import json
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT.parent / "Gold_Money_Data.xlsx"

panel = pd.read_csv(ROOT / "data/processed/panel_daily.csv", parse_dates=["date"])
gm = pd.read_csv(ROOT / "data/raw/lbma_gold_monthly.csv")
res = json.load(open(ROOT / "results/model_results.json"))
bt = json.load(open(ROOT / "results/backtest.json"))

wb = Workbook()
ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
TITLE = Font(name=ARIAL, bold=True, size=13, color="1F3864")
BOLD = Font(name=ARIAL, bold=True, size=10)
BASE = Font(name=ARIAL, size=10)


def style_header(ws, row, ncols):
    for j in range(1, ncols + 1):
        c = ws.cell(row=row, column=j)
        c.font = HDR_FONT
        c.fill = HDR_FILL


def write_df(ws, df, start=1, num_fmt=None):
    for i, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=start + i, column=j, value=v)
            c.font = BASE
            if num_fmt and i > 0 and j in num_fmt:
                c.number_format = num_fmt[j]
    style_header(ws, start, len(df.columns))


# ---------- ReadMe ----------
ws = wb.active
ws.title = "ReadMe"
rows = [
    ("Gold Money - data workbook", TITLE),
    ("", None),
    ("Companion to the gold-money GitHub project: do gold prices move the", BASE),
    ("currencies of gold-exporting countries? 21 years of daily evidence.", BASE),
    ("Built 2026-07-21. All series are values (a data extract, not a live model);", BASE),
    ("Summary_Stats contains live formulas over the panel.", BASE),
    ("", None),
    ("Sheets", BOLD),
    ("  Panel_Daily - merged daily panel 2005-01-04..2026-02-25 (5,146 rows)", BASE),
    ("  Summary_Stats - live formulas (AVERAGE / CORREL / MIN / MAX) on the panel", BASE),
    ("  Regressions - OLS results (JPY terms, USD terms, controls model)", BASE),
    ("  Tests - unit roots, cointegration, Granger causality, ECM, safe haven, OOS", BASE),
    ("  Backtest - momentum trading strategy results", BASE),
    ("  Gold_Monthly - LBMA monthly average gold, 1950-2026 (context)", BASE),
    ("", None),
    ("Sources", BOLD),
    ("  Gold PM fix (daily): LBMA, via github.com/unbalancedparentheses/forex-centuries", BASE),
    ("  FX daily (all local-per-USD): Fed H.10, via github.com/datasets/exchange-rates", BASE),
    ("  Broad USD index DTWEXBGS: FRED, via forex-centuries", BASE),
    ("  VIX: CBOE, via github.com/datasets/finance-vix", BASE),
    ("  Gold monthly: LBMA, via github.com/datasets/gold-prices", BASE),
    ("", None),
    ("Conventions", BOLD),
    ("  Crosses are JPY per unit of currency, e.g. AUDJPY = (JPY/USD)/(AUD/USD).", BASE),
    ("  gold_jpy = gold_usd * jpy_per_usd. Models use natural logs (see scripts/).", BASE),
]
for i, (txt, f) in enumerate(rows, 1):
    c = ws.cell(row=i, column=1, value=txt)
    if f:
        c.font = f
ws.column_dimensions["A"].width = 95

# ---------- Panel_Daily ----------
ws = wb.create_sheet("Panel_Daily")
p = panel.copy()
p["date"] = p["date"].dt.date
fmt = {1: "yyyy-mm-dd"}
for j in range(2, len(p.columns) + 1):
    fmt[j] = "#,##0.00"
write_df(ws, p, num_fmt=fmt)
ws.freeze_panes = "B2"
ws.column_dimensions["A"].width = 12
for j in range(2, len(p.columns) + 1):
    ws.column_dimensions[get_column_letter(j)].width = 11

# ---------- Summary_Stats (live formulas) ----------
ws = wb.create_sheet("Summary_Stats")
ws["A1"] = "Live summary statistics over Panel_Daily (recalculate on change)"
ws["A1"].font = TITLE
cols = list(panel.columns)
n = len(panel)


def col_ref(name):
    j = cols.index(name) + 1
    L = get_column_letter(j)
    return f"Panel_Daily!{L}2:{L}{n + 1}"


hdr = ["Series", "Mean", "Min", "Max", "Corr with gold_jpy"]
for j, h in enumerate(hdr, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(hdr))
series = ["gold_usd", "gold_jpy", "audjpy", "zarjpy", "cadjpy", "chfjpy", "vix", "usd_broad"]
for i, s in enumerate(series, 4):
    ws.cell(row=i, column=1, value=s).font = BASE
    ws.cell(row=i, column=2, value=f"=AVERAGE({col_ref(s)})")
    ws.cell(row=i, column=3, value=f"=MIN({col_ref(s)})")
    ws.cell(row=i, column=4, value=f"=MAX({col_ref(s)})")
    if s not in ("gold_jpy",):
        ws.cell(row=i, column=5, value=f"=CORREL({col_ref(s)},{col_ref('gold_jpy')})")
    for j in range(2, 6):
        ws.cell(row=i, column=j).number_format = "#,##0.000"
        ws.cell(row=i, column=j).font = BASE
ws.cell(row=13, column=1,
        value="Note: level correlations overstate relationships for trending series; "
              "see Regressions sheet for the log-log / sub-sample treatment.").font = BASE
for L, w in zip("ABCDE", [14, 14, 14, 14, 18]):
    ws.column_dimensions[L].width = w

# ---------- Regressions ----------
ws = wb.create_sheet("Regressions")
r = 1
ws.cell(row=r, column=1, value="OLS: log(FX in JPY) on log(gold in JPY), Newey-West t-stats").font = TITLE
r += 2
rows_ = [["Currency", "Period", "Gold beta", "t-stat", "R2", "N"]]
for c in ["aud", "zar", "cad", "chf"]:
    for k, lbl in [("full_2005_2026", "2005-2026"), ("pre_gfc_2005_2012", "2005-2012"),
                   ("taper_2013_2019", "2013-2019"), ("recent_2020_2026", "2020-2026")]:
        d = res["bivariate_jpy"][c][k]
        rows_.append([c.upper(), lbl, d["beta"], d["t"], d["r2"], d["n"]])
df = pd.DataFrame(rows_[1:], columns=rows_[0])
write_df(ws, df, start=r, num_fmt={3: "0.000", 4: "0.0", 5: "0.000"})
r += len(df) + 3
ws.cell(row=r, column=1, value="Same regression, both series in USD (the dollar effect)").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), res["bivariate_usd"][c]["beta"], res["bivariate_usd"][c]["r2"]]
                   for c in ["aud", "zar", "cad", "chf"]],
                  columns=["Currency", "Gold beta (USD terms)", "R2"])
write_df(ws, df, start=r, num_fmt={2: "0.000", 3: "0.000"})
r += len(df) + 3
ws.cell(row=r, column=1,
        value="Controls model 2006-2025: log FXJPY ~ log goldJPY + log broad USD + log VIX").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), d["beta_gold"], d["t_gold"], d["beta_usd"], d["beta_vix"], d["r2"]]
                   for c, d in res["controls_model"].items()],
                  columns=["Currency", "Gold beta", "t(gold)", "USD beta", "VIX beta", "R2"])
write_df(ws, df, start=r, num_fmt={2: "0.000", 3: "0.0", 4: "0.000", 5: "0.000", 6: "0.000"})
for L, w in zip("ABCDEF", [11, 22, 12, 12, 12, 10]):
    ws.column_dimensions[L].width = w

# ---------- Tests ----------
ws = wb.create_sheet("Tests")
r = 1
ws.cell(row=r, column=1, value="Unit roots (ADF p-values, log levels with trend / first differences)").font = TITLE
r += 2
df = pd.DataFrame([[k, v["adf_level_p"], v["adf_diff_p"]] for k, v in res["unit_roots"].items()],
                  columns=["Series", "Level p", "Diff p"])
write_df(ws, df, start=r, num_fmt={2: "0.000", 3: "0.0000"})
r += len(df) + 3
ws.cell(row=r, column=1, value="Engle-Granger cointegration (FX vs gold, JPY terms)").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), d["eg_stat"], d["p"], "no" if d["p"] > 0.05 else "yes"]
                   for c, d in res["engle_granger"].items()],
                  columns=["Currency", "EG stat", "p-value", "Cointegrated at 5%?"])
write_df(ws, df, start=r, num_fmt={2: "0.00", 3: "0.000"})
r += len(df) + 3
ws.cell(row=r, column=1, value="Johansen trace ({FX, gold, broad USD}), rank at 95%").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), str(d["trace_stats"]), str(d["cv_95"]), d["rank_95"]]
                   for c, d in res["johansen"].items()],
                  columns=["Currency", "Trace stats", "95% critical values", "Rank"])
write_df(ws, df, start=r)
r += len(df) + 3
ws.cell(row=r, column=1, value="Granger causality (daily returns, best of 5 lags, p-values)").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), d["gold_causes_fx_p"], d["fx_causes_gold_p"]]
                   for c, d in res["granger"].items()],
                  columns=["Currency", "Gold -> FX p", "FX -> Gold p"])
write_df(ws, df, start=r, num_fmt={2: "0.0000", 3: "0.0000"})
r += len(df) + 3
ws.cell(row=r, column=1, value="Error-correction term (indicative; no cointegration found)").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), d["ect_coef"], d["ect_t"], d["half_life_days"]]
                   for c, d in res["ecm"].items()],
                  columns=["Currency", "ECT coef", "t", "Half-life (days)"])
write_df(ws, df, start=r, num_fmt={2: "0.0000", 3: "0.0"})
r += len(df) + 3
ws.cell(row=r, column=1, value="Safe-haven check: daily return correlation with gold, by VIX regime").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), d["corr_all"], d["corr_low_vix"], d["corr_high_vix"]]
                   for c, d in res["safe_haven_corr"].items()],
                  columns=["Currency", "All days", "Low VIX (<median)", "High VIX (top decile)"])
write_df(ws, df, start=r, num_fmt={2: "0.000", 3: "0.000", 4: "0.000"})
r += len(df) + 3
ws.cell(row=r, column=1, value="Out-of-sample (realized gold, one-step, vs random walk)").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), d["mse_ratio_gold_vs_rw"], d["oos_r2_pct"]]
                   for c, d in res["out_of_sample"].items()],
                  columns=["Currency", "MSE ratio (model/RW)", "OOS R2 %"])
write_df(ws, df, start=r, num_fmt={2: "0.0000", 3: "0.00"})
for L, w in zip("ABCD", [13, 22, 22, 20]):
    ws.column_dimensions[L].width = w

# ---------- Backtest ----------
ws = wb.create_sheet("Backtest")
r = 1
ws.cell(row=r, column=1,
        value="Momentum-on-model-break strategy: 50d window, R2>=0.5, "
              "2-sigma entry, 10d hold, 2% stop; 2015-2026").font = TITLE
r += 2
df = pd.DataFrame([[c.upper(), bt[c]["n_trades"], bt[c]["total_ret_pct"], bt[c]["win_rate"],
                    bt[c]["avg_ret_pct"], bt[c]["best_pct"], bt[c]["worst_pct"]]
                   for c in ["aud", "zar", "cad", "chf"]],
                  columns=["Currency", "Trades", "Total ret %", "Win rate",
                           "Avg ret %", "Best %", "Worst %"])
write_df(ws, df, start=r, num_fmt={3: "0.00", 4: "0.00", 5: "0.000", 6: "0.00", 7: "0.00"})
r += len(df) + 3
ws.cell(row=r, column=1, value="AUD parameter grid: total return % (rows=holding days, cols=stop)").font = TITLE
r += 2
g = bt["aud_grid"]
ws.cell(row=r, column=1, value="hold \\ stop").font = HDR_FONT
ws.cell(row=r, column=1).fill = HDR_FILL
for j, s in enumerate(g["stops"], 2):
    c = ws.cell(row=r, column=j, value=s)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.number_format = "0.0%"
for i, h in enumerate(g["holds"]):
    ws.cell(row=r + 1 + i, column=1, value=h).font = BOLD
    for j, v in enumerate(g["total_ret_pct"][i], 2):
        c = ws.cell(row=r + 1 + i, column=j, value=v)
        c.number_format = "0.00"
        c.font = BASE
for L, w in zip("ABCDEFG", [13, 11, 12, 10, 10, 10, 10]):
    ws.column_dimensions[L].width = w

# ---------- Gold_Monthly ----------
ws = wb.create_sheet("Gold_Monthly")
gm.columns = ["Month", "Gold USD/oz (LBMA monthly avg)"]
write_df(ws, gm, num_fmt={2: "#,##0.00"})
ws.freeze_panes = "A2"
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 28

wb.save(OUT)
print("saved", OUT)
